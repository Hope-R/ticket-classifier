import os
import re

import pandas as pd


def process_tickets(run_month):

    # ==================================================
    # 0️⃣ RUN MONTH NORMALIZATION
    # ==================================================

    def normalize_run_month_text(value):
        """
        Normalize user-entered month text into the expected format: Jan-26
        without changing the underlying business rule.
        """
        if value is None:
            return ""

        value = str(value).strip()

        if not value:
            return ""

        value = value.replace("_", "-").replace("/", "-")
        value = re.sub(r"\s+", "", value)

        match = re.fullmatch(r"([A-Za-z]{3})-(\d{2})", value)
        if not match:
            return value

        mon, yy = match.groups()
        return f"{mon.title()}-{yy}"

    run_month = normalize_run_month_text(run_month)

    # ==================================================
    # 1️⃣ BASE PATHS
    # ==================================================

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    CONFIG_DIR = os.path.join(BASE_DIR, "config")
    RAW_DIR = os.path.join(BASE_DIR, "raw_data")
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")

    MONTH_RAW_DIR = os.path.join(RAW_DIR, run_month)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ==================================================
    # 2️⃣ VALIDATE INPUT MONTH FOLDER
    # ==================================================

    if not os.path.exists(MONTH_RAW_DIR):
        print(f"❌ Month folder not found: {MONTH_RAW_DIR}")
        return None

    def resolve_input_file(folder_path, base_name):
        """
        Allow either .xlsx or .csv for raw input files.
        Fail if none found or more than one found.
        """
        candidates = []

        xlsx_file = os.path.join(folder_path, f"{base_name}.xlsx")
        csv_file = os.path.join(folder_path, f"{base_name}.csv")

        if os.path.exists(xlsx_file):
            candidates.append(xlsx_file)

        if os.path.exists(csv_file):
            candidates.append(csv_file)

        if len(candidates) == 0:
            print(f"❌ Missing file for '{base_name}' in {folder_path}.")
            print(f"   Expected one of: {base_name}.xlsx or {base_name}.csv")
            return None

        if len(candidates) > 1:
            print(f"❌ Duplicate input files found for '{base_name}' in {folder_path}.")
            print("   Please keep only one of the following:")
            for c in candidates:
                print(f"   - {os.path.basename(c)}")
            return None

        return candidates[0]

    incident_file = resolve_input_file(MONTH_RAW_DIR, "incident")
    ur_file = resolve_input_file(MONTH_RAW_DIR, "ur")
    task_file = resolve_input_file(MONTH_RAW_DIR, "task")

    if not all([incident_file, ur_file, task_file]):
        return None

    # ==================================================
    # 2️⃣A CSV WARNING (WARNING-ONLY, NO PROMPT)
    # ==================================================

    if any(f.lower().endswith(".csv") for f in [incident_file, ur_file, task_file]):
        print("\n⚠️ Input format guidance:")
        print(".xlsx is the preferred raw input format.")
        print(".csv is supported, but depending on the source export, it may introduce formatting differences.")
        print("For the most consistent results, use .xlsx whenever available.\n")

    # ==================================================
    # 3️⃣ LOAD CONFIG FILES
    # ==================================================

    RULES_FILE = os.path.join(CONFIG_DIR, "business_rules.xlsx")

    if not os.path.exists(RULES_FILE):
        print(f"❌ Rules file not found: {RULES_FILE}")
        return None

    def read_excel_safe(file_path, sheet_name=0, dtype=None, keep_default_na=True):
        """
        Read Excel more safely across environments/pandas versions.
        Default to first sheet unless a sheet name is explicitly passed.
        """
        try:
            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                dtype=dtype,
                keep_default_na=keep_default_na,
                engine="openpyxl",
            )
        except TypeError:
            return pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                dtype=dtype,
                engine="openpyxl",
            )
        except ImportError:
            print("❌ openpyxl is required to read Excel files. Please install it with: pip install openpyxl")
            return None
        except Exception as e:
            print(f"❌ Failed to read Excel file: {file_path}")
            if sheet_name is not None:
                print(f"   Sheet: {sheet_name}")
            print(f"   Error: {e}")
            return None

    assignment_groups_df = read_excel_safe(
        RULES_FILE,
        sheet_name="assignment_groups"
    )
    keyword_rules_df = read_excel_safe(
        RULES_FILE,
        sheet_name="keyword_rules"
    )
    phrase_rules_df = read_excel_safe(
        RULES_FILE,
        sheet_name="phrase_rules"
    )
    template_noise_df = read_excel_safe(
        RULES_FILE,
        sheet_name="template_noise"
    )
    category_rules_df = read_excel_safe(
        RULES_FILE,
        sheet_name="category_check"
    )
    category_subcategory_rules_df = read_excel_safe(
        RULES_FILE,
        sheet_name="category_subcategory_check"
    )
    column_mapping_df = read_excel_safe(
        RULES_FILE,
        sheet_name="column_mapping"
    )
    final_category_consolidation_df = read_excel_safe(
        RULES_FILE,
        sheet_name="final_category_consolidation"
    )

    config_dfs = [
        assignment_groups_df,
        keyword_rules_df,
        phrase_rules_df,
        template_noise_df,
        category_rules_df,
        category_subcategory_rules_df,
        column_mapping_df,
        final_category_consolidation_df,
    ]

    if any(df is None for df in config_dfs):
        return None

    # ==================================================
    # 4️⃣ COLUMN STANDARDIZATION
    # ==================================================

    column_mapping_df.columns = column_mapping_df.columns.str.strip()

    column_mapping_df["Standard Column"] = (
        column_mapping_df["Standard Column"].astype(str).str.strip()
    )

    column_mapping_df["Possible Column Name"] = (
        column_mapping_df["Possible Column Name"]
        .astype(str)
        .str.strip()
        .str.lower()
    )

    column_aliases = (
        column_mapping_df
        .groupby("Standard Column", sort=True)["Possible Column Name"]
        .apply(list)
        .to_dict()
    )
    print("\n🔎 OPENED ALIASES FROM BUSINESS_RULES:")
    print(column_aliases.get("Opened"))

    required_columns = [
        "Number", "Caller", "Email", "Contact type", "Opened",
        "Short description", "Description", "Assignment group",
        "Priority", "Category", "Sub category", "Service",
        "Resolved", "Close notes"
    ]

    logic_required_columns = required_columns + ["Primary Ticket"]

    def standardize_columns(df):
        df = df.copy()
        df.columns = [str(col).strip() for col in df.columns]
        lower_cols = {col.lower(): col for col in df.columns}

        rename_dict = {}

        for standard_col, aliases in column_aliases.items():
            for alias in aliases:
                if alias in lower_cols:
                    rename_dict[lower_cols[alias]] = standard_col

        df = df.rename(columns=rename_dict)

        for col in logic_required_columns:
            if col not in df.columns:
                df[col] = ""

        return df.loc[:, ~df.columns.duplicated()]

    # ==================================================
    # 5️⃣ LOAD RAW FILES (CANONICAL INGESTION)
    # ==================================================

    def normalize_invisible_chars(text):
        """
        Normalize invisible / special whitespace characters and Excel-escaped
        line break artifacts that may differ between Excel and CSV exports.
        """
        if pd.isna(text):
            return ""

        text = str(text)

        text = re.sub(r"_x000D_", "\r", text, flags=re.IGNORECASE)
        text = re.sub(r"_x000A_", "\n", text, flags=re.IGNORECASE)
        text = re.sub(r"_x0009_", "\t", text, flags=re.IGNORECASE)

        replacements = {
            "\u00A0": " ",
            "\u2007": " ",
            "\u202F": " ",
            "\u200B": "",
            "\u200C": "",
            "\u200D": "",
            "\u2060": "",
            "\uFEFF": "",
        }

        for bad, good in replacements.items():
            text = text.replace(bad, good)

        return text

    def canonicalize_scalar(value):
        """
        Canonical standard:
        - Treat blank-like values consistently
        - Preserve text content
        - Strip leading/trailing spaces
        - Normalize hidden/invisible characters
        """
        if pd.isna(value):
            return ""

        if isinstance(value, str):
            value = normalize_invisible_chars(value)
            value = value.strip()

            if value.lower() in {"nan", "none", "null"}:
                return ""

            return value

        return value

    def load_raw_file(file_path):
        """
        Canonical ingestion layer:
        - Read .xlsx and .csv in a controlled way
        - Avoid pandas guessing types differently across formats
        - Standardize blanks/strings before business logic runs
        """
        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == ".xlsx":
                df = read_excel_safe(
                    file_path,
                    sheet_name=0,
                    dtype=str,
                    keep_default_na=False
                )
                if df is None:
                    return None
            elif ext == ".csv":
                try:
                    df = pd.read_csv(
                        file_path,
                        dtype=str,
                        keep_default_na=False,
                        encoding="utf-8-sig"
                    )
                except UnicodeDecodeError:
                    df = pd.read_csv(
                        file_path,
                        dtype=str,
                        keep_default_na=False,
                        encoding="latin1"
                    )
            else:
                print(f"❌ Unsupported file format: {file_path}")
                return None
        except Exception as e:
            print(f"❌ Failed to load raw file: {file_path}")
            print(f"   Error: {e}")
            return None

        df = df.apply(lambda col: col.map(canonicalize_scalar))
        return df

    incident_df = load_raw_file(incident_file)
    ur_df = load_raw_file(ur_file)
    task_df = load_raw_file(task_file)
    if task_df is not None:
        print("\n🔎 TASK CSV RAW COLUMNS AFTER LOAD:")
        for col in task_df.columns:
            print(repr(col))

        print("\n🔎 Does raw Task file already contain exact 'Opened'?")
        print("Opened" in task_df.columns)

    if any(df is None for df in [incident_df, ur_df, task_df]):
        return None

    incident_df = standardize_columns(incident_df)
    ur_df = standardize_columns(ur_df)
    task_df = standardize_columns(task_df)

    # ==================================================
    # 6️⃣ BUSINESS FILTERS
    # ==================================================

    incident_df = incident_df[
        incident_df["Contact type"].astype(str).str.strip() != "Alert"
    ]

    ur_df = ur_df[
        (
            ur_df["Primary Ticket"].isna() |
            (ur_df["Primary Ticket"].astype(str).str.strip() == "")
        )
        &
        (
            ur_df["Category"].notna() &
            (ur_df["Category"].astype(str).str.strip() != "")
        )
    ]

    valid_groups = (
        assignment_groups_df.iloc[:, 0]
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )

    task_df = task_df[
        task_df["Assignment group"].astype(str).str.strip().isin(valid_groups)
    ]

    # ==================================================
    # 7️⃣ ADD METADATA
    # ==================================================

    for df, ttype in [(incident_df, "Inc"), (ur_df, "UR"), (task_df, "Task")]:
        df["Ticket Type"] = ttype

    # ==================================================
    # 8️⃣ COMBINE
    # ==================================================

    end_user_tickets = pd.concat(
        [incident_df, ur_df, task_df],
        ignore_index=True
    )

    # ==================================================
    # 9️⃣ TEXT NORMALIZATION / CANONICAL HELPERS
    # ==================================================

    def preserve_display_text(text):
        if pd.isna(text):
            return ""

        text = normalize_invisible_chars(text)
        text = str(text)
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        text = text.strip()

        if text.lower() in {"nan", "none", "null"}:
            return ""

        return text

    def normalize_matching_text(text):
        if pd.isna(text):
            return ""

        text = normalize_invisible_chars(text)
        text = str(text).strip()
        text = re.sub(r"\s+", " ", text)
        text = text.lower()
        return text

    def normalize_canonical_key(text):
        if pd.isna(text):
            return ""

        text = normalize_invisible_chars(text)
        text = str(text).strip()
        text = re.sub(r"\s+", " ", text)
        text = text.lower()
        return text

    if "Short description" in end_user_tickets.columns:
        end_user_tickets["Short description"] = (
            end_user_tickets["Short description"].apply(preserve_display_text)
        )
        end_user_tickets["Short description_canonical"] = (
            end_user_tickets["Short description"].apply(normalize_matching_text)
        )
    else:
        end_user_tickets["Short description"] = ""
        end_user_tickets["Short description_canonical"] = ""

    if "Description" in end_user_tickets.columns:
        end_user_tickets["Description"] = (
            end_user_tickets["Description"].apply(preserve_display_text)
        )
        end_user_tickets["Description_canonical"] = (
            end_user_tickets["Description"].apply(normalize_matching_text)
        )
    else:
        end_user_tickets["Description"] = ""
        end_user_tickets["Description_canonical"] = ""

    if "Close notes" in end_user_tickets.columns:
        end_user_tickets["Close notes"] = (
            end_user_tickets["Close notes"].apply(preserve_display_text)
        )
    else:
        end_user_tickets["Close notes"] = ""

    if "Service" in end_user_tickets.columns:
        end_user_tickets["Service"] = (
            end_user_tickets["Service"].apply(preserve_display_text)
        )
        end_user_tickets["Service_canonical"] = (
            end_user_tickets["Service"].apply(normalize_canonical_key)
        )
    else:
        end_user_tickets["Service"] = ""
        end_user_tickets["Service_canonical"] = ""

    if "Category" in end_user_tickets.columns:
        end_user_tickets["Category"] = (
            end_user_tickets["Category"].apply(preserve_display_text)
        )
        end_user_tickets["Category_canonical"] = (
            end_user_tickets["Category"].apply(normalize_canonical_key)
        )
    else:
        end_user_tickets["Category"] = ""
        end_user_tickets["Category_canonical"] = ""

    if "Sub category" in end_user_tickets.columns:
        end_user_tickets["Sub category"] = (
            end_user_tickets["Sub category"].apply(preserve_display_text)
        )
        end_user_tickets["Sub category_canonical"] = (
            end_user_tickets["Sub category"].apply(normalize_canonical_key)
        )
    else:
        end_user_tickets["Sub category"] = ""
        end_user_tickets["Sub category_canonical"] = ""

    # ==================================================
    # 🔟 DERIVE MONTH FROM OPENED COLUMN
    # ==================================================

    end_user_tickets["Opened"] = pd.to_datetime(
        end_user_tickets["Opened"],
        errors="coerce"
    )

    end_user_tickets["Month"] = end_user_tickets["Opened"].dt.strftime("%b-%y")

    month_values = (
        end_user_tickets["Month"]
        .dropna()
        .astype(str)
        .str.strip()
        .unique()
        .tolist()
    )

    if not month_values:
        print("❌ Could not derive Month from 'Opened' column. Please check input data.")
        return None

    if len(month_values) > 1:
        print(f"❌ Multiple months detected in input data: {', '.join(sorted(month_values))}")
        print(f"❌ Selected month folder: {run_month}")
        print("❌ Please ensure only one month of ticket data is placed in the selected folder.")
        return None

    derived_month = month_values[0]

    if derived_month != run_month:
        print("❌ Mismatch detected.")
        print(f"❌ Selected month folder: {run_month}")
        print(f"❌ Derived month from 'Opened': {derived_month}")
        print("❌ Please check the raw files and selected month folder.")
        return None

    # ==================================================
    # 1️⃣1️⃣ PREPARE RULE LOGIC
    # ==================================================

    keyword_rules_df.columns = keyword_rules_df.columns.str.strip()
    keyword_rules_df["Category"] = (
        keyword_rules_df["Category"].astype(str).str.strip()
    )
    keyword_rules_df["Keyword"] = (
        keyword_rules_df["Keyword"].astype(str).str.lower().str.strip()
    )

    if "Search_Field" in keyword_rules_df.columns:
        keyword_rules_df["Search_Field"] = (
            keyword_rules_df["Search_Field"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .replace({
                "short-description": "short_description",
                "short description": "short_description",
                "description": "description",
                "both": "both",
                "": "both"
            })
        )
        keyword_rules_df.loc[
            ~keyword_rules_df["Search_Field"].isin(["short_description", "description", "both"]),
            "Search_Field"
        ] = "both"
    else:
        keyword_rules_df["Search_Field"] = "both"

    keyword_rules_df["Priority"] = pd.to_numeric(
        keyword_rules_df["Priority"],
        errors="coerce"
    ).fillna(999)

    keyword_rules_df = keyword_rules_df.sort_values(
        by=["Category", "Keyword", "Priority", "Search_Field"],
        kind="stable"
    ).reset_index(drop=True)

    category_keywords = {}
    category_priority = {}
    category_keyword_search_fields = {}

    for _, row in keyword_rules_df.iterrows():
        category = row["Category"]
        keyword = row["Keyword"]
        priority = int(row["Priority"])
        search_field = row["Search_Field"]

        category_keywords.setdefault(category, []).append(keyword)
        category_priority[category] = min(
            priority,
            category_priority.get(category, 999)
        )
        category_keyword_search_fields[(category, keyword)] = search_field

    for category in category_keywords:
        category_keywords[category] = sorted(set(category_keywords[category]))

    phrase_rules_df.columns = phrase_rules_df.columns.str.strip()
    phrase_rules_df["Category"] = (
        phrase_rules_df["Category"].astype(str).str.strip()
    )
    phrase_rules_df["Phrase"] = (
        phrase_rules_df["Phrase"].astype(str).str.lower().str.strip()
    )
    phrase_rules_df["Weight"] = pd.to_numeric(
        phrase_rules_df["Weight"],
        errors="coerce"
    ).fillna(0)

    if "Search_Field" in phrase_rules_df.columns:
        phrase_rules_df["Search_Field"] = (
            phrase_rules_df["Search_Field"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.lower()
            .replace({
                "short-description": "short_description",
                "short description": "short_description",
                "description": "description",
                "both": "both",
                "": "both"
            })
        )
        phrase_rules_df.loc[
            ~phrase_rules_df["Search_Field"].isin(["short_description", "description", "both"]),
            "Search_Field"
        ] = "both"
    else:
        phrase_rules_df["Search_Field"] = "both"

    phrase_rules_df = phrase_rules_df.sort_values(
        by=["Category", "Phrase", "Weight", "Search_Field"],
        kind="stable"
    ).reset_index(drop=True)

    category_phrases = {}
    category_phrase_search_fields = {}

    for _, row in phrase_rules_df.iterrows():
        category = row["Category"]
        phrase = row["Phrase"]
        weight = row["Weight"]
        search_field = row["Search_Field"]

        category_phrases.setdefault(category, []).append((phrase, weight))
        category_phrase_search_fields[(category, phrase)] = search_field

    for category in category_phrases:
        category_phrases[category] = sorted(
            category_phrases[category],
            key=lambda x: (x[0], x[1])
        )

    # ==================================================
    # 1️⃣2️⃣ LOAD CATEGORY & CATEGORY+SUBCATEGORY RULES
    # ==================================================

    category_rules_df["Category"] = (
        category_rules_df["Category"].astype(str).str.lower().str.strip()
    )
    category_rules_df["Mapped Category"] = (
        category_rules_df["Mapped Category"].astype(str).str.strip()
    )

    category_rules_df = category_rules_df.sort_values(
        by=["Category", "Mapped Category"],
        kind="stable"
    ).reset_index(drop=True)

    category_rules = dict(
        zip(category_rules_df["Category"], category_rules_df["Mapped Category"])
    )

    category_subcategory_rules_df.columns = (
        category_subcategory_rules_df.columns.str.strip()
    )

    category_subcategory_rules_df["Category"] = (
        category_subcategory_rules_df["Category"]
        .fillna("")
        .astype(str)
        .apply(normalize_canonical_key)
    )

    category_subcategory_rules_df["Sub category"] = (
        category_subcategory_rules_df["Sub category"]
        .fillna("")
        .astype(str)
        .apply(normalize_canonical_key)
    )

    category_subcategory_rules_df["Mapped Category"] = (
        category_subcategory_rules_df["Mapped Category"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    category_subcategory_rules_df = category_subcategory_rules_df[
        (category_subcategory_rules_df["Category"] != "") &
        (category_subcategory_rules_df["Sub category"] != "") &
        (category_subcategory_rules_df["Mapped Category"] != "")
    ].copy()

    category_subcategory_rules_df = category_subcategory_rules_df.sort_values(
        by=["Category", "Sub category", "Mapped Category"],
        kind="stable"
    ).reset_index(drop=True)

    category_subcategory_map = dict(
        zip(
            zip(
                category_subcategory_rules_df["Category"],
                category_subcategory_rules_df["Sub category"]
            ),
            category_subcategory_rules_df["Mapped Category"]
        )
    )

    # ==================================================
    # 1️⃣3️⃣ LOAD FINAL CATEGORY CONSOLIDATION RULES
    # ==================================================

    final_category_consolidation_df.columns = (
        final_category_consolidation_df.columns.str.strip()
    )

    final_category_consolidation_df["Source Value"] = (
        final_category_consolidation_df["Source Value"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    final_category_consolidation_df["Final Category"] = (
        final_category_consolidation_df["Final Category"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    final_category_consolidation_df = final_category_consolidation_df[
        (final_category_consolidation_df["Source Value"] != "") &
        (final_category_consolidation_df["Final Category"] != "")
    ].copy()

    final_category_consolidation_df["Source Value_canonical"] = (
        final_category_consolidation_df["Source Value"].apply(normalize_canonical_key)
    )

    final_category_consolidation_df = final_category_consolidation_df.sort_values(
        by=["Source Value_canonical", "Final Category"],
        kind="stable"
    ).reset_index(drop=True)

    final_category_map = dict(
        zip(
            final_category_consolidation_df["Source Value_canonical"],
            final_category_consolidation_df["Final Category"]
        )
    )

    # ==================================================
    # 1️⃣4️⃣ TEMPLATE SANITIZATION
    # ==================================================

    template_noise_df["Phrase"] = (
        template_noise_df["Phrase"]
        .astype(str)
        .apply(normalize_invisible_chars)
        .str.strip()
    )

    template_noise_df = template_noise_df.sort_values(
        by=["Phrase"],
        kind="stable"
    ).reset_index(drop=True)

    TEMPLATE_PHRASES = (
        template_noise_df["Phrase"]
        .dropna()
        .tolist()
    )

    def clean_description(text):
        cleaned = text

        for phrase in TEMPLATE_PHRASES:
            pattern = re.compile(re.escape(phrase), re.IGNORECASE)
            cleaned = pattern.sub("", cleaned)

        return cleaned

    end_user_tickets["Description_clean"] = (
        end_user_tickets["Description_canonical"].apply(clean_description)
    )

    # ==================================================
    # 1️⃣5️⃣ WEIGHTED SCORING ENGINE
    # ==================================================

    MIN_SCORE_THRESHOLD = 1

    def get_text_by_search_field(short_desc_text, description_text, combined_text, search_field):
        if search_field == "short_description":
            return short_desc_text
        if search_field == "description":
            return description_text
        return combined_text

    def determine_category(row):

        category_value = str(row.get("Category_canonical", "")).strip()
        subcategory_value = str(row.get("Sub category_canonical", "")).strip()
        service_value = str(row.get("Service_canonical", "")).strip()

        short_desc = str(row.get("Short description_canonical", "")).strip()
        description = str(row.get("Description_clean", "")).strip()
        combined_text = (short_desc + " " + description).strip()

        # 1) Category + Sub category targeted check
        if category_value and subcategory_value:
            category_subcategory_key = (category_value, subcategory_value)
            if category_subcategory_key in category_subcategory_map:
                return category_subcategory_map[category_subcategory_key]

        # 2) Category override
        if category_value in category_rules:
            return category_rules[category_value]

        # 3) Service fallback
        if service_value:
            return str(row.get("Service", "")).strip()

        # 4) Microsoft Teams special rule
        if "Microsoft Teams" in category_keywords:
            for keyword in sorted(category_keywords["Microsoft Teams"]):
                if keyword in short_desc:
                    return "Microsoft Teams"

        # 5) Weighted scoring
        scores = {}

        all_categories = sorted(
            set(category_keywords.keys()) | set(category_phrases.keys())
        )

        for category in all_categories:

            if category == "Microsoft Teams":
                continue

            phrase_score = 0
            keyword_score = 0

            for phrase, weight in category_phrases.get(category, []):
                search_field = category_phrase_search_fields.get(
                    (category, phrase),
                    "both"
                )
                text_to_search = get_text_by_search_field(
                    short_desc,
                    description,
                    combined_text,
                    search_field
                )
                if phrase in text_to_search:
                    phrase_score += weight

            for keyword in category_keywords.get(category, []):
                search_field = category_keyword_search_fields.get(
                    (category, keyword),
                    "both"
                )
                text_to_search = get_text_by_search_field(
                    short_desc,
                    description,
                    combined_text,
                    search_field
                )
                pattern = r"\b" + re.escape(keyword) + r"\b"
                if re.search(pattern, text_to_search):
                    keyword_score += 1

            total_score = phrase_score + keyword_score

            if total_score > 0:
                scores[category] = {
                    "total": total_score,
                    "phrase": phrase_score,
                    "priority": category_priority.get(category, 999)
                }

        # 6) Others fallback
        if not scores:
            return "Others"

        sorted_categories = sorted(
            scores.items(),
            key=lambda x: (
                -x[1]["total"],
                -x[1]["phrase"],
                x[1]["priority"],
                x[0]
            )
        )

        best_category, best_data = sorted_categories[0]

        if best_data["total"] >= MIN_SCORE_THRESHOLD:
            return best_category

        return "Others"

    end_user_tickets["Ticket Category"] = (
        end_user_tickets.apply(determine_category, axis=1)
    )

    # ==================================================
    # 1️⃣6️⃣ FINAL CATEGORY CONSOLIDATION
    # ==================================================

    end_user_tickets["Ticket Category"] = (
        end_user_tickets["Ticket Category"]
        .astype(str)
        .str.strip()
    )

    end_user_tickets["Ticket Category"] = (
        end_user_tickets["Ticket Category"]
        .apply(lambda x: final_category_map.get(normalize_canonical_key(x), x))
    )

    # ==================================================
    # 1️⃣7️⃣ FINAL OUTPUT PREPARATION
    # ==================================================

    final_columns = required_columns + [
        "Ticket Type", "Month", "Ticket Category"
    ]

    end_user_tickets = end_user_tickets[final_columns].copy()

    end_user_tickets["Resolved"] = pd.to_datetime(
        end_user_tickets["Resolved"],
        errors="coerce"
    )

    def remove_garbage_rows(df):
        if not isinstance(df, pd.DataFrame):
            print("⚠️ Skipping invalid object during cleanup; expected DataFrame.")
            return pd.DataFrame()

        df = df.copy()

        required_cleanup_columns = ["Number", "Opened", "Short description", "Description"]

        for col in required_cleanup_columns:
            if col not in df.columns:
                df[col] = ""

        df = df[
            ~df["Number"].astype(str).str.contains(
                "Export stopped", case=False, na=False
            )
        ].copy()

        df = df.dropna(how="all")

        df = df[
            ~(
                (df["Number"].astype(str).str.strip() == "") &
                (df["Opened"].astype(str).str.strip() == "") &
                (df["Short description"].astype(str).str.strip() == "") &
                (df["Description"].astype(str).str.strip() == "")
            )
        ].copy()

        return df

    def sanitize_excel_cell(value):
        if pd.isna(value):
            return value

        if not isinstance(value, str):
            return value

        text = value

        if len(text) > 32767:
            text = text[:32767]

        if text.startswith(("=", "+", "-", "@")):
            text = "'" + text

        return text

    def sanitize_for_excel(df):
        df = df.copy()
        text_columns = df.select_dtypes(include=["object"]).columns.tolist()

        for col in text_columns:
            df[col] = df[col].apply(sanitize_excel_cell)

        return df

    def apply_datetime_format_to_excel(file_path, datetime_columns, datetime_format="yyyy-mm-dd hh:mm:ss"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("❌ openpyxl is required to format Excel files. Please install it with: pip install openpyxl")
            return False
        except Exception as e:
            print(f"⚠️ Could not import openpyxl for formatting: {e}")
            return False

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            header_to_col_idx = {}
            for col_idx, cell in enumerate(ws[1], start=1):
                header_to_col_idx[str(cell.value).strip()] = col_idx

            for col_name in datetime_columns:
                col_idx = header_to_col_idx.get(col_name)
                if not col_idx:
                    continue

                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and cell.value != "":
                        cell.number_format = datetime_format

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"⚠️ Could not apply Excel datetime formatting to {file_path}")
            print(f"   Error: {e}")
            return False

    end_user_tickets = remove_garbage_rows(end_user_tickets)

    output_file = os.path.join(
        OUTPUT_DIR,
        f"end_user_ticket_data_{run_month.lower()}.xlsx"
    )

    master_file = os.path.join(
        OUTPUT_DIR,
        "master_ticket_data.xlsx"
    )

    # ==================================================
    # 1️⃣8️⃣ SAVE MONTHLY OUTPUT
    # ==================================================

    monthly_output_df = sanitize_for_excel(end_user_tickets.copy())

    try:
        monthly_output_df.to_excel(output_file, index=False, engine="openpyxl")
    except ImportError:
        print("❌ openpyxl is required to write Excel files. Please install it with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ Failed to save monthly output: {output_file}")
        print(f"   Error: {e}")
        return None

    monthly_datetime_format_applied = apply_datetime_format_to_excel(
        output_file,
        ["Opened", "Resolved"]
    )

    # ==================================================
    # 1️⃣9️⃣ REBUILD MASTER FILE FROM MONTHLY OUTPUTS
    # ==================================================

    monthly_files = []

    try:
        for file_name in os.listdir(OUTPUT_DIR):
            if (
                file_name.startswith("end_user_ticket_data_")
                and file_name.endswith(".xlsx")
            ):
                monthly_files.append(os.path.join(OUTPUT_DIR, file_name))
    except Exception as e:
        print(f"❌ Failed to scan output directory: {OUTPUT_DIR}")
        print(f"   Error: {e}")
        return None

    monthly_files = sorted(monthly_files)

    master_parts = []

    for monthly_file in monthly_files:
        df = read_excel_safe(monthly_file, sheet_name=0)
        if df is None:
            print(f"⚠️ Skipping unreadable monthly output file: {monthly_file}")
            continue

        if "Opened" in df.columns:
            df["Opened"] = pd.to_datetime(df["Opened"], errors="coerce")

        if "Resolved" in df.columns:
            df["Resolved"] = pd.to_datetime(df["Resolved"], errors="coerce")

        df = remove_garbage_rows(df)

        if df.empty:
            print(f"⚠️ Skipping empty or invalid monthly output file: {monthly_file}")
            continue

        master_parts.append(df)

    if master_parts:
        master_df = pd.concat(master_parts, ignore_index=True)
    else:
        master_df = pd.DataFrame(columns=end_user_tickets.columns)

    if "Number" in master_df.columns:
        master_df = master_df.drop_duplicates(subset=["Number"], keep="last")

    master_df = remove_garbage_rows(master_df)

    master_output_df = sanitize_for_excel(master_df.copy())

    try:
        master_output_df.to_excel(master_file, index=False, engine="openpyxl")
    except ImportError:
        print("❌ openpyxl is required to write Excel files. Please install it with: pip install openpyxl")
        return None
    except Exception as e:
        print(f"❌ Failed to save master output: {master_file}")
        print(f"   Error: {e}")
        return None

    master_datetime_format_applied = apply_datetime_format_to_excel(
        master_file,
        ["Opened", "Resolved"]
    )

    print("✅ Weighted scoring processing complete.")
    print("✅ Final category consolidation applied.")
    print("✅ Month derived from 'Opened' column.")
    print("✅ Strict month validation passed.")
    print("✅ Excel-safe export sanitization applied.")
    print("✅ Raw input supports .xlsx and .csv.")
    print("✅ Canonical ingestion applied.")
    print("✅ Original text preserved in output.")
    print("✅ Canonical helper fields used for matching logic.")
    print("✅ Deterministic rule ordering applied.")
    print("✅ Hidden character normalization applied.")
    print("✅ Category + Sub category targeted mapping applied when configured.")
    print("✅ Service fallback uses raw Service value when needed.")
    print("✅ Search_Field support applied to keyword and phrase rules.")
    print("✅ Final category consolidation uses canonical in-memory matching.")
    print("✅ 'Resolved' exported as datetime.")
    if monthly_datetime_format_applied and master_datetime_format_applied:
        print("✅ Excel date-time formatting applied to 'Opened' and 'Resolved'.")
    else:
        print("⚠️ Excel date-time formatting could not be fully applied, but files were still saved.")
    print("✅ Monthly output saved to:", output_file)
    print("✅ Master file rebuilt from monthly outputs:", master_file)

    return end_user_tickets